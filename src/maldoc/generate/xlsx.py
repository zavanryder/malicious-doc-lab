"""XLSX document generation with openpyxl."""

from pathlib import Path

from maldoc.attacks.base import AttackResult


def generate_xlsx(attack_result: AttackResult, title: str, output_path: Path) -> Path:
    """Generate an XLSX workbook containing the attack payload."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Document"

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Content"
    ws["B2"] = "Notes"

    row = 3
    for line in attack_result.visible_content.splitlines():
        cleaned = line.strip()
        if cleaned:
            ws.cell(row=row, column=1, value=cleaned)
            row += 1

    if attack_result.metadata:
        props = wb.properties
        props.creator = attack_result.metadata.get("author", "")
        props.subject = attack_result.metadata.get("subject", "")
        props.keywords = attack_result.metadata.get("keywords", "")
        props.description = attack_result.metadata.get("description", "")

    hints = attack_result.format_hints or {}
    if attack_result.technique == "white_on_white":
        cell = ws.cell(row=max(3, row), column=5, value=attack_result.hidden_content)
        cell.font = Font(color="FFFFFF", size=1)
        ws.column_dimensions["E"].hidden = True
    elif attack_result.technique == "off_page":
        # Place data far to the right/bottom where many viewers won't show by default,
        # without using extreme sheet bounds that can cause pathological scan times.
        ws.cell(row=5000, column=200, value=attack_result.hidden_content)
    elif hints.get("obfuscated_variants"):
        ws.cell(row=max(3, row), column=2, value=hints["obfuscated_variants"][0])
    elif hints.get("obfuscated_payload"):
        ws.cell(row=max(3, row), column=2, value=hints["obfuscated_payload"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
